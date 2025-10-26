import pandas as pd
import re
import sqlite3
import networkx as nx
from openpyxl import load_workbook

uploaded_file = r"C:\Users\ADMIN\Desktop\AI_XL_Agent\data\valuation_test_case.xlsx"

sheets = pd.read_excel(uploaded_file, sheet_name=None) 

def is_wide_format(df: pd.DataFrame) -> bool:
    time_like = re.compile(r"^\d{4}$|Q\d+-\d{4}", re.IGNORECASE)
    return sum(bool(time_like.match(str(c))) for c in df.columns) >= (len(df.columns) / 2)

def reshape_wide(df: pd.DataFrame) -> pd.DataFrame:
    return df.melt(id_vars=[df.columns[0]], var_name="Time", value_name="Value")

# normalized_sheets = {}
# for sheet, df in sheets.items():
#     normalized_sheets[sheet] = reshape_wide(df) if is_wide_format(df) else df
    
# conn = sqlite3.connect("workbook.db")

# for sheet, df in normalized_sheets.items():
#     df.to_sql(sheet, conn, if_exists="replace", index=False)

# wb = load_workbook(uploaded_file, data_only=False)
# graph = nx.DiGraph()

# for sheet in wb.sheetnames:
#     ws = wb[sheet]
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.data_type == "f":  # formula cell
#                 graph.add_node(f"{sheet}!{cell.coordinate}", formula=cell.value)



